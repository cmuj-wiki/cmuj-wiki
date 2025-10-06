#!/usr/bin/env python3
"""
COLOR-AWARE Schedule Parser
Handles Excel schedules where classes are grouped by background color
"""

import openpyxl
import re
import json
import sys
from pathlib import Path

DAY_MAP = {
    'Poniedziałek': 0,
    'Wtorek': 1,
    'Środa': 2,
    'Czwartek': 3,
    'Piątek': 4,
}

def get_cell_color(cell):
    """Extract fill color from cell (returns string identifier)"""
    try:
        if not cell.fill or cell.fill.fill_type != 'solid':
            return None
        
        fg_color = cell.fill.fgColor
        if not fg_color:
            return None
        
        # Try to get a consistent color identifier
        if hasattr(fg_color, 'rgb') and fg_color.rgb:
            return str(fg_color.rgb)
        elif hasattr(fg_color, 'theme') and fg_color.theme is not None:
            return f"theme{fg_color.theme}"
        elif hasattr(fg_color, 'index') and fg_color.index:
            return f"index{fg_color.index}"
        
        return None
    except:
        return None

def find_class_text_for_color(ws, column, target_row, target_color):
    """
    Look up in the same column to find a cell with the same color that has text.
    This handles the Excel pattern where only the first group has text,
    but all groups share the same background color.
    """
    if not target_color:
        return None
    
    # Search upward from target row
    for row_num in range(target_row - 1, 0, -1):
        cell = ws.cell(row=row_num, column=column)
        cell_color = get_cell_color(cell)
        
        if cell_color == target_color and cell.value:
            return cell.value
    
    return None

print("🎨 COLOR-AWARE PARSER - Starting...")
print("=" * 70)

wb = openpyxl.load_workbook('I-rok-2024-2025_www_zimowy.xlsx')

all_classes = []

for sheet_name in wb.sheetnames:
    if 'english' in sheet_name.lower():
        continue
    
    day_name = sheet_name.strip()
    if day_name not in DAY_MAP:
        continue
    
    day_num = DAY_MAP[day_name]
    ws = wb[sheet_name]
    
    print(f"\n📅 {day_name}")
    
    # Find header row
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(), 1):
        if row[1].value == 'h':
            header_row_idx = i
            break
    
    if not header_row_idx:
        continue
    
    # Parse each group row
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        group_cell = row[1]  # Column B
        if not group_cell.value:
            continue
        
        group_text = str(group_cell.value).strip()
        if not group_text.startswith('gr.'):
            continue
        
        group_match = re.search(r'gr\.\s*(\d+)', group_text)
        if not group_match:
            continue
        
        group_num = int(group_match.group(1))
        row_num = group_cell.row
        
        # Iterate through time slot columns (C onwards)
        for cell in row[2:]:  # Skip columns A, B
            # Skip if no value AND no color
            cell_color = get_cell_color(cell)
            
            if not cell.value and not cell_color:
                continue
            
            # Get the class name
            class_text = cell.value
            
            # If cell is colored but empty, look up for the text
            if not class_text and cell_color:
                class_text = find_class_text_for_color(ws, cell.column, row_num, cell_color)
            
            if not class_text:
                continue
            
            # Calculate time
            col_offset = cell.column - 3  # Column C = 0
            hour = 8 + (col_offset // 4)
            minute = (col_offset % 4) * 15
            
            # Simple duration: 90 min default
            duration = 90
            
            print(f"  Group {group_num}: {str(class_text)[:30]} at {hour:02d}:{minute:02d}")
            
            all_classes.append({
                'group': group_num,
                'day': day_num,
                'day_name': day_name,
                'hour': hour,
                'minute': minute,
                'duration': duration,
                'subject': str(class_text).strip(),
                'type': None,
                'location': None,
                'dates': [],
                'time': None
            })

print(f"\n✅ Found {len(all_classes)} classes")

# Save
output_dir = Path('docs/static')
output_dir.mkdir(exist_ok=True)

with open(output_dir / 'schedule_test.json', 'w', encoding='utf-8') as f:
    json.dump(all_classes, f, ensure_ascii=False, indent=2)

print(f"💾 Saved to {output_dir / 'schedule_test.json'}")

# Test: Group 8 Tuesday
group_8_tue = [c for c in all_classes if c['group'] == 8 and c['day'] == 1]
print(f"\n📊 Group 8 Tuesday classes: {len(group_8_tue)}")
for cls in group_8_tue:
    print(f"  - {cls['subject']}")

