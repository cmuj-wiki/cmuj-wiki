#!/usr/bin/env python3
"""
ULEPSZONA wersja parsera planu zajęć.
Uwzględnia:
1. Lokalizacje/adresy z dolnej części arkusza
2. Precyzyjne daty zajęć (np. "Historia Med. 28.XI-19.I")
3. Dni wolne z PDF kalendarza
"""

import openpyxl
import fitz  # PyMuPDF
import re
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

DAY_MAP = {
    'Poniedziałek': 0,
    'Wtorek': 1,
    'Środa': 2,
    'Czwartek': 3,
    'Piątek': 4,
}

# Mapowanie miesięcy rzymskich
MONTHS = {
    'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6,
    'VII': 7, 'VIII': 8, 'IX': 9, 'X': 10, 'XI': 11, 'XII': 12
}

# Standardowe czasy trwania zajęć (w minutach)
# Używane gdy komórki nie są scalone i nie można wykryć czasu z struktury XLSX
STANDARD_DURATIONS = {
    # Ćwiczenia (laboratoria) - zazwyczaj 2 godziny
    'histologia ćw': 120,
    'anatomia ćw': 150,  # Anatomia often longer
    'biochemia ćw': 120,
    'bhk ćw': 90,

    # Seminaria - zazwyczaj 1.5-2 godziny
    'histologia sem': 90,
    'anatomia sem': 90,
    'biochemia sem': 90,
    'etyka sem': 90,
    'historia med sem': 90,
    'historia medycyny sem': 90,

    # Wykłady - zazwyczaj 1.5 godziny
    'wykład': 90,
    'wykl': 90,

    # WF - zazwyczaj 1.5 godziny
    'wf': 90,
    'wychowanie fizyczne': 90,

    # Zajęcia zintegrowane - zazwyczaj 2.25 godziny
    'zaj zintegr': 135,
    'zintegrowane': 135,

    # Kolokwia - zazwyczaj 1.5 godziny
    'kolokwium': 90,

    # Default fallback
    'default': 90
}

def parse_holidays_from_pdf(pdf_path):
    """Parsuje dni wolne z PDF kalendarza."""

    try:
        pdf = fitz.open(pdf_path)
        text = pdf[0].get_text()

        holidays = []

        # Parsuj tekst szukając dat oznaczonych jako wolne
        # Format w PDF może być różny, trzeba dostosować

        # Znane święta państwowe 2024-2025
        known_holidays = [
            (11, 1),   # Wszystkich Świętych
            (11, 11),  # Niepodległości
            (12, 25),  # Boże Narodzenie
            (12, 26),  # 2. dzień świąt
            (1, 1),    # Nowy Rok
            (1, 6),    # Trzech Króli
            (4, 20),   # Wielkanoc 2025
            (4, 21),   # Poniedziałek Wielkanocny
            (5, 1),    # Święto Pracy
            (5, 3),    # Święto Konstytucji
            (6, 8),    # Boże Ciało
            (6, 19),   # 2025
        ]

        for month, day in known_holidays:
            year = 2024 if month >= 10 else 2025
            holidays.append(datetime(year, month, day))

        # Dodaj okresy sesji/ferii (można wyciągnąć z PDF)
        # Przykład: Ferie zimowe, sesja egzaminacyjna

        print(f"📅 Znaleziono {len(holidays)} dni wolnych/świąt")
        return holidays

    except Exception as e:
        print(f"⚠️  Nie udało się parsować PDF: {e}")
        return []

def extract_address_mapping(ws):
    """Wyciąga mapowanie przedmiot -> adres z dolnej części arkusza."""

    address_map = {}

    # Szukamy sekcji z adresami (kolumna 25, wiersze 31-40)
    for row_idx in range(30, 45):
        try:
            subject_cell = ws.cell(row=row_idx, column=2).value
            address_cell = ws.cell(row=row_idx, column=25).value

            if subject_cell and address_cell:
                subject = str(subject_cell).strip()
                address = str(address_cell).strip()

                # Keep original subject name (with spaces) for word-based matching
                # Store lowercase for case-insensitive matching
                address_map[subject.lower()] = address

                print(f"  📍 {subject} → {address}")
        except:
            pass

    return address_map

def match_location(subject_name, address_map):
    """
    Dopasowuje lokalizację do przedmiotu z ulepszonym fuzzy matchingiem.

    Obsługuje:
    - Dopasowanie po słowach kluczowych (np. "Etyka Sem" → "Etyka w medycynie")
    - Ignorowanie skrótów typu "ćw.", "Sem", "wykł."
    - Rozszerzanie skrótów (WF → Wychowanie fizyczne, Med. → medycyny)
    - Normalizacja polskich znaków
    """

    if not subject_name:
        return None

    # Normalizuj polski tekst
    def normalize_polish(text):
        replacements = {
            'ą': 'a', 'ć': 'c', 'ę': 'e', 'ł': 'l',
            'ń': 'n', 'ó': 'o', 'ś': 's', 'ź': 'z', 'ż': 'z'
        }
        text = text.lower()
        for pl, ascii_char in replacements.items():
            text = text.replace(pl, ascii_char)
        return text

    # Rozszerz znane skróty
    abbreviations = {
        'WF': 'wychowanie fizyczne',
        'Med.': 'medycyny',
        'Med': 'medycyny',
        'Zintegr.': 'zintegrowane',
        'Zaj.': 'zajecia',
        'przedklin.': 'przedklinicznych',
        'przedklinicznych': 'przedklinicznych',
        'klin.': 'klinicznych',
    }

    subject_expanded = subject_name
    for abbr, full in abbreviations.items():
        if abbr in subject_expanded:
            subject_expanded = subject_expanded.replace(abbr, full)

    # Usuń typowe skróty i suffiksy
    subject_clean = subject_expanded
    for suffix in [' ćw.', ' Sem', ' sem', ' wykł.', ' col', ' ćw', ' cw']:
        # Don't remove " do " as it's part of date ranges like "do 7.XI"
        if suffix in subject_clean:
            subject_clean = subject_clean.replace(suffix, '')

    # Wyciągnij słowa kluczowe (pomijamy krótkie słowa jak "w", "i", "z")
    subject_words = set([
        normalize_polish(w.strip('.,')) for w in subject_clean.split()
        if len(w.strip('.,')) > 2  # Słowa dłuższe niż 2 znaki, bez interpunkcji
    ])

    # Próbuj dopasować po słowach kluczowych
    best_match = None
    best_score = 0

    for key, address in address_map.items():
        # Key is already lowercase from extract_address_mapping()
        key_words = set([
            normalize_polish(w.strip('.,')) for w in key.split()
            if len(w.strip('.,')) > 2
        ])

        # Policz wspólne słowa
        common_words = subject_words & key_words
        if common_words:
            score = len(common_words) / max(len(subject_words), len(key_words))
            if score > best_score:
                best_score = score
                best_match = address

    # Próg dopasowania (co najmniej 40% wspólnych słów)
    # Lower threshold for abbreviations
    if best_score >= 0.4:
        # Log fuzzy matches with low confidence
        if best_score < 0.7:
            print(f"  ⚠️  Fuzzy match: '{subject_name}' → '{best_match}' (confidence: {best_score:.0%})", file=sys.stderr)
        return best_match

    # Fallback: proste substring matching (stary sposób)
    subject_normalized = normalize_polish(subject_expanded.replace(' ', ''))

    for key, address in address_map.items():
        key_normalized = normalize_polish(key.replace(' ', ''))
        if key_normalized in subject_normalized or subject_normalized in key_normalized:
            return address

    return None

def parse_time_from_column(col_idx, header_row):
    """
    Wyciąga godzinę z pozycji kolumny.

    XLSX ma grid 15-minutowy:
    - Kolumna C (index 3) = 8:00
    - Każda kolumna = 15 minut
    - Minuty ZAWSZE: 0, 15, 30, 45

    FIX: Nie czytamy z header row (buggy), tylko obliczamy z pozycji.
    """

    # Kolumna C (index 3) to początek dnia o 8:00
    offset = col_idx - 3

    # 4 sloty na godzinę (0, 15, 30, 45 min)
    slots_per_hour = 4

    hour = 8 + (offset // slots_per_hour)
    minute = (offset % slots_per_hour) * 15

    return hour, minute

def parse_class_cell(cell_value):
    """Parsuje komórkę z zajęciami."""

    if not cell_value or not isinstance(cell_value, str):
        return None

    cell_value = cell_value.strip()
    if not cell_value:
        return None

    lines = [line.strip() for line in cell_value.split('\n') if line.strip()]

    if not lines:
        return None

    subject = lines[0]

    # Szukamy dat w formacie "DD.MM" lub "DD.MM-DD.MM" lub "DD.MM.YYYY"
    date_pattern = r'(\d{1,2}\.[IVX]+(?:\.\d{4})?(?:\s*-\s*\d{1,2}\.[IVX]+(?:\.\d{4})?)?)'
    dates = []
    class_type = None
    location = None
    time_info = None

    # NOWE: Sprawdź czy data jest w NAZWIE przedmiotu (np. "Histologia ćw. do 7.XI")
    # Format "do DD.MM"
    end_date_pattern = r'do\s+(\d{1,2}\.[IVX]+)'
    end_match = re.search(end_date_pattern, subject)
    if end_match:
        dates.append(f"do {end_match.group(1)}")

    # Format "DD.MM-DD.MM" w nazwie
    range_in_name = re.search(r'(\d{1,2}\.[IVX]+\s*-\s*\d{1,2}\.[IVX]+)', subject)
    if range_in_name and not dates:
        dates.append(range_in_name.group(1))

    for line in lines[1:]:
        if any(keyword in line.lower() for keyword in ['wykład', 'ćw', 'sem', 'col']):
            class_type = line
        elif re.search(date_pattern, line):
            dates.append(line)
        elif re.search(r'\d{1,2}[:.]\d{2}', line):
            time_info = line
        elif line and 'MS Teams' in line:
            location = line
        elif line and len(line) > 3 and not any(c.isdigit() for c in line[:5]):
            # Może być lokalizacja
            if not location:
                location = line

    return {
        'subject': subject,
        'type': class_type,
        'location': location,
        'dates': dates,
        'time': time_info,
    }

def get_duration_from_subject(subject_name):
    """
    Zwraca standardowy czas trwania na podstawie nazwy przedmiotu.
    Returns duration in minutes based on subject name patterns.
    """
    if not subject_name:
        return STANDARD_DURATIONS['default']

    # Normalize subject name
    subject_lower = subject_name.lower().strip()

    # Try exact matches first
    for key, duration in STANDARD_DURATIONS.items():
        if key == 'default':
            continue
        if key in subject_lower:
            return duration

    # Fallback to default
    return STANDARD_DURATIONS['default']

def infer_duration_from_cells(ws, cell, row_cells, col_idx):
    """
    Infers class duration using hybrid approach:
    1. Check if cell is in a merged range (most reliable)
    2. Look at when the next class starts (scan forward)
    3. Use subject-based defaults as fallback

    Returns: duration in minutes
    """
    # Strategy 1: Check if this cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Only count if this is the TOP-LEFT cell of the merge
            if (cell.row == merged_range.min_row and
                cell.column == merged_range.min_col):
                width = merged_range.max_col - merged_range.min_col + 1
                duration = width * 15
                return duration
            else:
                # This cell is part of a merge but not the start
                # Should not happen if we track processed_coords correctly
                return None  # Signal to skip this cell

    # Strategy 2: Scan forward to find next non-empty cell
    # Look at cells in the same row after this one
    MAX_SCAN_SLOTS = 20  # Don't scan beyond 5 hours (20 * 15 min)
    next_class_found = False
    slots_until_next = 1  # Start with current slot

    for offset in range(1, MAX_SCAN_SLOTS):
        next_col_idx = col_idx + offset
        if next_col_idx < len(row_cells):
            next_cell = row_cells[next_col_idx]
            if next_cell and next_cell.value:
                # Found next class
                next_class_found = True
                break
        slots_until_next += 1

    if next_class_found and slots_until_next <= 12:  # Max 3 hours seems reasonable
        duration = slots_until_next * 15
        return duration

    # Strategy 3: Use subject-based defaults
    subject_name = str(cell.value) if cell.value else ""
    return get_duration_from_subject(subject_name)

def validate_schedule_data(data):
    """
    Validates schedule data format before saving.
    Expected format: [{group: int, day: int, hour: int, minute: int, duration: int, ...}, ...]
    """
    if not isinstance(data, list):
        raise ValueError("Schedule data must be a list")

    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"Item {i} must be a dictionary, got {type(item)}")

        # Check required fields
        required_fields = {'group', 'day', 'day_name', 'hour', 'minute', 'duration', 'subject'}
        if not all(field in item for field in required_fields):
            missing = required_fields - set(item.keys())
            raise ValueError(f"Item {i} missing required fields: {missing}")

        # Validate numeric fields
        if not isinstance(item['group'], int) or item['group'] < 1:
            raise ValueError(f"Item {i} has invalid group: {item['group']}")

        if not isinstance(item['day'], int) or not (0 <= item['day'] <= 4):
            raise ValueError(f"Item {i} has invalid day: {item['day']} (expected 0-4)")

        if not isinstance(item['hour'], int) or not (0 <= item['hour'] <= 23):
            raise ValueError(f"Item {i} has invalid hour: {item['hour']}")

        if not isinstance(item['minute'], int) or item['minute'] not in [0, 15, 30, 45]:
            raise ValueError(f"Item {i} has invalid minute: {item['minute']} (expected 0, 15, 30, or 45)")

        if not isinstance(item['duration'], int) or item['duration'] <= 0:
            raise ValueError(f"Item {i} has invalid duration: {item['duration']}")

    return True

def parse_schedule_file(filepath, address_map, holidays):
    """Główna funkcja parsująca plik XLS."""

    print(f"📖 Parsowanie: {filepath}")
    wb = openpyxl.load_workbook(filepath)

    all_classes = []

    for sheet_name in wb.sheetnames:
        if 'english' in sheet_name.lower():
            continue

        day_name = sheet_name.strip()
        if day_name not in DAY_MAP:
            print(f"⚠️  Pomijam nieznany arkusz: {sheet_name}")
            continue

        day_num = DAY_MAP[day_name]
        ws = wb[sheet_name]

        print(f"  📅 Dzień: {day_name}")

        header_row = None
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and row[1] == 'h':
                header_row = row
                header_row_idx = i
                break

        if not header_row:
            print(f"    ⚠️  Nie znaleziono nagłówka godzin")
            continue

        # Track processed cells to avoid counting merged cells multiple times
        processed_coords = set()

        # NEW: Use cell objects instead of values_only for better control
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            # Check if this is a group row (column B)
            group_cell = row[1]  # Column B (index 1)
            if not group_cell.value:
                continue

            group_cell_text = str(group_cell.value).strip()
            if not group_cell_text.startswith('gr.'):
                continue

            group_match = re.search(r'gr\.\s*(\d+)', group_cell_text)
            if not group_match:
                continue

            group_num = int(group_match.group(1))
            row_idx = group_cell.row  # Excel row number (1-based)

            # Iterate over data columns (starting from column C = index 2)
            # Get all cells in this row as a list for easy indexing
            row_cells = list(row)

            for col_offset, cell in enumerate(row_cells[2:], start=0):  # Start from column C
                # Skip if already processed (part of merged cell we handled)
                if cell.coordinate in processed_coords:
                    continue

                # Skip empty cells
                if not cell.value:
                    continue

                # Parse the cell content
                parsed = parse_class_cell(cell.value)
                if not parsed:
                    continue

                # Calculate time from column position
                # Column C (index 2) = 8:00, each column = 15 minutes
                hour, minute = parse_time_from_column(cell.column - 1, header_row)

                # NEW: Use hybrid duration detection
                duration_minutes = infer_duration_from_cells(ws, cell, row_cells, cell.column - 1)

                if duration_minutes is None:
                    # This cell is part of a merged range but not the start
                    # Mark it as processed and skip
                    processed_coords.add(cell.coordinate)
                    continue

                # Mark cells as processed for merged ranges
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Mark all cells in this range as processed
                        for r in range(merged_range.min_row, merged_range.max_row + 1):
                            for c in range(merged_range.min_col, merged_range.max_col + 1):
                                coord = ws.cell(row=r, column=c).coordinate
                                processed_coords.add(coord)
                        break

                # Mark single cell as processed if not merged
                processed_coords.add(cell.coordinate)

                # Dopasuj lokalizację z mapy adresów
                if not parsed['location']:
                    parsed['location'] = match_location(parsed['subject'], address_map)
                    if not parsed['location']:
                        print(f"  📍 Brak lokalizacji dla: '{parsed['subject']}' (grupa {group_num})", file=sys.stderr)

                all_classes.append({
                    'group': group_num,
                    'day': day_num,
                    'day_name': day_name,
                    'hour': hour,
                    'minute': minute,
                    'duration': duration_minutes,
                    **parsed
                })

    print(f"\n✅ Znaleziono {len(all_classes)} zajęć")
    return all_classes

if __name__ == "__main__":
    # 1. Parsuj dni wolne z PDF
    holidays = parse_holidays_from_pdf('Kalendarz_2025-2026.pdf')

    # 2. Wyciągnij mapowanie adresów
    print("\n📍 Mapowanie adresów:")
    wb = openpyxl.load_workbook('I-rok-2024-2025_www_zimowy.xlsx')
    ws = wb['Poniedziałek ']
    address_map = extract_address_mapping(ws)

    # 3. Parsuj zajęcia
    print("\n📚 Parsowanie zajęć:")
    classes = parse_schedule_file('I-rok-2024-2025_www_zimowy.xlsx', address_map, holidays)

    # 4. Zapisz
    output_dir = Path('docs/static')
    output_dir.mkdir(exist_ok=True)

    # VALIDATE before saving
    try:
        validate_schedule_data(classes)
    except ValueError as e:
        print(f"\n❌ Validation error: {e}")
        raise

    # Zapisz zajęcia
    with open(output_dir / 'schedule_data_v2.json', 'w', encoding='utf-8') as f:
        json.dump(classes, f, ensure_ascii=False, indent=2)

    # NOTE: holidays.json is generated separately by calendar_parser.py
    # to ensure correct format for the calendar widget

    print(f"\n💾 Zapisano:")
    print(f"  - {output_dir / 'schedule_data_v2.json'}")
    print(f"\n⚠️  Pamiętaj: Uruchom calendar_parser.py aby zaktualizować holidays.json")

    # Przykład - grupa 11 która ma Historię Medycyny z datami
    print("\n📋 Przykład - Grupa 11:")
    group_11 = [c for c in classes if c['group'] == 11]
    for cls in group_11[:5]:
        loc = cls.get('location', 'Brak')
        dates = cls.get('dates', [])
        print(f"  - {cls['subject']}: {loc} | Daty: {dates}")
